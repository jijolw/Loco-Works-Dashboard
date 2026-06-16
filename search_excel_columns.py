import os
import openpyxl

def check_excel_file(path, out_file):
    out_file.write(f"\nChecking: {path}\n")
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        out_file.write(f"  Sheets: {wb.sheetnames}\n")
        for name in wb.sheetnames:
            ws = wb[name]
            rows = []
            # Read first 5 rows to ensure we get headers
            for r in ws.iter_rows(max_row=5, values_only=True):
                if any(r):
                    rows.append(r)
            if rows:
                out_file.write(f"    Sheet '{name}' headers:\n")
                out_file.write(f"      {str(rows[0][:20])}\n")
                # Search for keywords
                for row in rows:
                    for val in row:
                        if val and any(k in str(val).lower() for k in ["month", "return", "premature", "or", "interval", "outturn"]):
                            out_file.write(f"      -> Found keyword '{val}' in sheet '{name}'\n")
    except Exception as e:
        out_file.write(f"  Error reading {path}: {str(e)}\n")

def main():
    dirs = [r"C:\Users\User\Downloads", r"D:\JIJO\TARGET", r"D:\JIJO"]
    with open("excel_search_results.txt", "w", encoding="utf-8") as f:
        for d in dirs:
            if not os.path.exists(d):
                continue
            f.write(f"\nScanning directory: {d}\n")
            for file in os.listdir(d):
                if file.endswith((".xlsx", ".xlsm")):
                    path = os.path.join(d, file)
                    check_excel_file(path, f)

if __name__ == "__main__":
    main()
